"""Sell-side급 재무 모델 xlsx 생성기 — 함수 자동 연결 버전.

JPM CJ제일제당 + 삼성 실적추정 sell-side 표준 참고.
모든 합계·비율·교차 시트 참조가 formula로 작동.

구조 (10 시트):
  Cover / Summary / IS / BS / CF / Drivers / Segment_Buildup / Valuation_Band / SOTP / Notes

색 코드:
  🟡 노란 셀 = 사용자 input
  🟢 녹색 = 자동 formula (편집 가능)
"""
from __future__ import annotations

import logging
import re
import tempfile
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)


# ─────────────────────── 스타일 ───────────────────────
HEADER_FONT = Font(name="Malgun Gothic", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="0A3D3A")
SUBHEADER_FONT = Font(name="Malgun Gothic", size=10, bold=True, color="1A237E")
SUBHEADER_FILL = PatternFill("solid", fgColor="E8EAF6")
LABEL_FONT = Font(name="Malgun Gothic", size=10, bold=True)
DATA_FONT = Font(name="Malgun Gothic", size=10)
INPUT_FONT = Font(name="Malgun Gothic", size=10, color="1976D2", bold=True)
INPUT_FILL = PatternFill("solid", fgColor="FFF9C4")
FORMULA_FONT = Font(name="Malgun Gothic", size=10, color="2E7D32")
THIN = Side(border_style="thin", color="BBBBBB")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)


# ─────────────────────── 산업 템플릿 ───────────────────────
@dataclass
class IndustryTemplate:
    code: str
    label: str
    segments: list[str]
    drivers: list[tuple[str, str, str]]   # (segment, driver_name, unit)
    valuation_multiples: list[str]
    notes: str


INDUSTRY_TEMPLATES: dict[str, IndustryTemplate] = {
    "auto_parts": IndustryTemplate(
        code="auto_parts",
        label="자동차 부품",
        segments=["자동차부품", "방산", "기타"],
        drivers=[
            ("자동차부품", "글로벌 자동차 생산대수 (M units)", "M units"),
            ("자동차부품", "당사 점유율 (%)", "%"),
            ("자동차부품", "차량당 부품 수 (units)", "units"),
            ("자동차부품", "부품당 ASP (KRW)", "₩"),
            ("자동차부품", "GP 마진 (%)", "%"),
            ("자동차부품", "OP 마진 (%)", "%"),
            ("방산", "수주 잔고 (₩bn)", "₩bn"),
            ("방산", "연간 인식 비율 (%)", "%"),
            ("방산", "GP 마진 (%)", "%"),
            ("방산", "OP 마진 (%)", "%"),
            ("기타", "매출 (₩bn)", "₩bn"),
            ("기타", "OP 마진 (%)", "%"),
        ],
        valuation_multiples=["P/E", "P/B", "EV/EBITDA", "EV/Sales"],
        notes=("자동차부품: 현대차/기아 글로벌 생산 사이클 + EV 전환 영향.\n"
               "방산: 수주 잔고 기반 매출 가시성. K2 소총·권총 수출.\n"
               "라이트 capex 사업 → FCF 양호 → 순현금 포지션 확인."),
    ),
    "semis": IndustryTemplate(
        code="semis", label="반도체",
        segments=["DRAM", "NAND", "Foundry", "기타"],
        drivers=[
            ("DRAM", "B/G QoQ (%)", "%"), ("DRAM", "ASP QoQ (%)", "%"),
            ("NAND", "B/G QoQ (%)", "%"), ("NAND", "ASP QoQ (%)", "%"),
            ("Foundry", "Utilization (%)", "%"),
            ("Foundry", "Wafer ASP ($)", "$"),
        ],
        valuation_multiples=["P/E", "P/B", "EV/EBITDA"],
        notes="Memory cycle + HBM mix + capex 동향.",
    ),
    "saas": IndustryTemplate(
        code="saas", label="SaaS",
        segments=["Subscription", "Services"],
        drivers=[
            ("Subscription", "Net New ARR ($M)", "$M"),
            ("Subscription", "NRR (%)", "%"),
            ("Subscription", "Gross Margin (%)", "%"),
            ("Services", "Revenue ($M)", "$M"),
        ],
        valuation_multiples=["EV/Revenue", "EV/ARR", "P/FCF"],
        notes="Rule of 40, magic number, S&M efficiency.",
    ),
    "biotech": IndustryTemplate(
        code="biotech", label="Biotech (commercial)",
        segments=["Product1", "Product2", "Pipeline"],
        drivers=[
            ("Product1", "Patients (K)", "K"),
            ("Product1", "Price per year ($K)", "$K"),
            ("Product1", "Penetration (%)", "%"),
            ("Product1", "GP 마진 (%)", "%"),
            ("Pipeline", "R&D ($M)", "$M"),
            ("Pipeline", "Probability of success (%)", "%"),
        ],
        valuation_multiples=["EV/Revenue", "EV/Peak Sales", "P/E"],
        notes="Cash runway = cash / quarterly burn. 임상 readout 카탈리스트.",
    ),
    "cpg": IndustryTemplate(
        code="cpg", label="Consumer / Food",
        segments=["Food", "Bio", "Logistics", "기타"],
        drivers=[
            ("Food", "Volume growth (%)", "%"),
            ("Food", "ASP growth (%)", "%"),
            ("Food", "GP 마진 (%)", "%"),
            ("Bio", "Lysine price ($/ton)", "$/ton"),
            ("Bio", "Methionine price ($/ton)", "$/ton"),
            ("Bio", "Volume (kton)", "kton"),
        ],
        valuation_multiples=["P/E", "EV/EBITDA", "SOTP"],
        notes="Raw material spread 모니터링.",
    ),
    "general": IndustryTemplate(
        code="general", label="General",
        segments=["Core"],
        drivers=[
            ("Core", "Revenue growth (%)", "%"),
            ("Core", "GP 마진 (%)", "%"),
            ("Core", "OP 마진 (%)", "%"),
        ],
        valuation_multiples=["P/E", "EV/EBITDA"],
        notes="기본 driver only.",
    ),
}


# ─────────────────────── 헬퍼 ───────────────────────
def _set_widths(ws, widths: dict[int, int]):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _hdr(ws, row: int, col: int, text: str, span: int = 1):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)


def _subhdr(ws, row: int, col: int, text: str, span: int = 1):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = SUBHEADER_FONT
    cell.fill = SUBHEADER_FILL
    cell.alignment = LEFT
    cell.border = BORDER
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)


def _label(ws, row: int, col: int, text: str, indent: int = 0):
    cell = ws.cell(row=row, column=col, value=("  " * indent) + text)
    cell.font = LABEL_FONT
    cell.alignment = LEFT
    cell.border = BORDER


def _input_cell(ws, row: int, col: int, value=None, fmt: str = "#,##0.00"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = INPUT_FONT
    cell.fill = INPUT_FILL
    cell.alignment = RIGHT
    cell.border = BORDER
    if fmt:
        cell.number_format = fmt


def _formula_cell(ws, row: int, col: int, formula: str, fmt: str = "#,##0.0"):
    cell = ws.cell(row=row, column=col, value=formula)
    cell.font = FORMULA_FONT
    cell.alignment = RIGHT
    cell.border = BORDER
    if fmt:
        cell.number_format = fmt


def _periods(hist_years, proj_years):
    return [f"{y}A" for y in hist_years] + [f"{y}E" for y in proj_years]


# ─────────────────────── Formula resolver ───────────────────────
def _resolve(expr: str, col_letter: str, row_map: dict[str, int],
             cross_refs: dict[str, dict[str, int]] | None = None) -> str:
    """expr 안의 [label] / [sheet:label] 패턴 → 실제 셀 참조로 변환.
    SUM([a]:[b])도 동작 — [a]와 [b]가 같은 col_letter의 다른 row가 됨."""
    cross_refs = cross_refs or {}

    def repl(m):
        ref = m.group(1).strip()
        if ":" in ref and not ref.startswith(" "):
            # sheet:label 형식
            parts = ref.split(":", 1)
            sheet, label = parts[0].strip(), parts[1].strip()
            if sheet in cross_refs and label in cross_refs[sheet]:
                r = cross_refs[sheet][label]
                # 시트명 공백 포함 시 quote
                sheet_quoted = f"'{sheet}'" if " " in sheet or "_" in sheet else sheet
                return f"{sheet_quoted}!{col_letter}{r}"
            return ref
        # 같은 시트
        if ref in row_map:
            return f"{col_letter}{row_map[ref]}"
        return ref

    return re.sub(r"\[([^\[\]]+)\]", repl, expr)


def _build_sheet(
    ws,
    spec: list[dict],
    label_col: int,
    data_col_start: int,
    n_periods: int,
    cross_refs: dict[str, dict[str, int]] | None = None,
    row_start: int = 3,
) -> dict[str, int]:
    """단일 시트 빌드 — spec 기반.
    spec = [{label, kind, expr?, fmt?, indent?}, ...]
    kind = 'input' | 'formula' | 'subheader' | 'blank' | 'header'
    expr = "[Revenue] - [(-) COGS]" 같은 표현식 ([label]은 같은 시트 내 row 참조)
    """
    cross_refs = cross_refs or {}
    row_map: dict[str, int] = {}
    row = row_start
    for item in spec:
        kind = item.get("kind", "input")
        label = item.get("label", "")
        if kind == "blank":
            row += 1
            continue
        if kind == "header":
            _hdr(ws, row, label_col, label, span=1 + n_periods)
            row += 1
            continue
        if kind == "subheader":
            _subhdr(ws, row, label_col, label, span=1 + n_periods)
            row += 1
            continue
        # input / formula
        _label(ws, row, label_col, label, indent=item.get("indent", 0))
        if label:
            row_map[label] = row
        fmt = item.get("fmt", "#,##0.0")
        for i in range(n_periods):
            col = data_col_start + i
            col_letter = get_column_letter(col)
            if kind == "input":
                _input_cell(ws, row, col, value=item.get("value"), fmt=fmt)
            elif kind == "formula":
                expr = item.get("expr", "")
                if expr:
                    resolved = _resolve(expr, col_letter, row_map, cross_refs)
                    _formula_cell(ws, row, col, f"={resolved}", fmt=fmt)
                else:
                    _formula_cell(ws, row, col, "", fmt=fmt)
        row += 1
    return row_map


# ─────────────────────── 시트 빌더 ───────────────────────
def _build_cover(wb, ticker, name, industry_label, analyst):
    ws = wb.create_sheet("Cover", 0)
    _set_widths(ws, {1: 4, 2: 38, 3: 60})
    ws.row_dimensions[3].height = 30
    ws.merge_cells("B3:C3")
    h = ws.cell(row=3, column=2, value=f"{name} ({ticker})")
    h.font = Font(name="Malgun Gothic", size=18, bold=True, color="0A3D3A")
    h.alignment = LEFT

    rows = [
        ("Industry template", industry_label),
        ("Ticker", ticker),
        ("Company", name),
        ("Analyst", analyst or "—"),
        ("Model date", "=TEXT(TODAY(),\"yyyy-mm-dd\")"),
        ("", ""),
        ("Color code",
         "🟡 노란 셀 = 사용자 input · 🟢 녹색 = 자동 formula · 일반 = label/header"),
        ("Convention",
         "단위: 한국 ₩bn / 미국 $M 기본. 분기 = '1Q25', 연간 = '2025A/E'."),
        ("Workflow",
         "1) Drivers 입력 → 2) Segment_Buildup 채움 → 3) IS/BS/CF 자동 합산 → "
         "4) Valuation_Band/SOTP에서 implied price → 5) Summary 한 페이지."),
        ("Reference",
         "JPM CJ제일제당 모델 (24 시트) + 삼성전자 실적추정 (driver chain) 표준 참고"),
    ]
    for i, (k, v) in enumerate(rows, start=5):
        _label(ws, i, 2, k)
        c = ws.cell(row=i, column=3, value=v)
        c.font = DATA_FONT
        c.alignment = LEFT
        c.border = BORDER


def _build_drivers(wb, template: IndustryTemplate, n_periods: int,
                   periods: list[str]) -> dict[str, int]:
    ws = wb.create_sheet("Drivers")
    _set_widths(ws, {1: 4, 2: 22, 3: 38, 4: 10})
    for c in range(5, 5 + n_periods):
        ws.column_dimensions[get_column_letter(c)].width = 11

    _hdr(ws, 1, 2, f"📊 Drivers — {template.label}", span=3 + n_periods)
    _subhdr(ws, 2, 2, "Segment")
    _subhdr(ws, 2, 3, "Driver")
    _subhdr(ws, 2, 4, "Unit")
    for i, p in enumerate(periods):
        _subhdr(ws, 2, 5 + i, p)

    row_map: dict[str, int] = {}
    row = 3
    for segment, driver, unit in template.drivers:
        _label(ws, row, 2, segment)
        _label(ws, row, 3, driver)
        c = ws.cell(row=row, column=4, value=unit)
        c.font = DATA_FONT
        c.alignment = CENTER
        c.border = BORDER
        for i in range(n_periods):
            fmt = "0.00%" if "%" in driver else "#,##0.0"
            _input_cell(ws, row, 5 + i, fmt=fmt)
        # row_map은 driver 이름으로 (Segment_Buildup에서 참조용)
        row_map[f"{segment}::{driver}"] = row
        row += 1

    # 가이드 노트
    row += 1
    _subhdr(ws, row, 2, "💡 가이드", span=3 + n_periods)
    row += 1
    for line in template.notes.split("\n"):
        c = ws.cell(row=row, column=2, value=line)
        c.font = DATA_FONT
        c.alignment = WRAP
        ws.merge_cells(start_row=row, start_column=2,
                       end_row=row, end_column=4 + n_periods)
        row += 1
    return row_map


def _build_segment(wb, template: IndustryTemplate, n_periods: int,
                   periods: list[str]) -> dict[str, int]:
    """Segment_Buildup — 각 segment별 Revenue/GP/OP/OPM, 마지막에 Total = SUM."""
    ws = wb.create_sheet("Segment_Buildup")
    _set_widths(ws, {1: 4, 2: 22, 3: 18})
    for c in range(4, 4 + n_periods):
        ws.column_dimensions[get_column_letter(c)].width = 12

    _hdr(ws, 1, 2, f"🧩 Segment Buildup — {template.label}", span=2 + n_periods)
    _subhdr(ws, 2, 2, "Segment")
    _subhdr(ws, 2, 3, "Line item")
    for i, p in enumerate(periods):
        _subhdr(ws, 2, 4 + i, p)

    # 각 segment마다 4줄 (Revenue/GP/OP/OPM), 1줄 공백
    row_map: dict[str, int] = {}
    row = 3
    segment_revenue_rows: list[int] = []
    segment_gp_rows: list[int] = []
    segment_op_rows: list[int] = []

    for segment in template.segments:
        _label(ws, row, 2, segment)
        # Revenue, GP, OP — input
        for line, fmt in [("Revenue", "#,##0.0"), ("GP", "#,##0.0"),
                          ("OP", "#,##0.0")]:
            _label(ws, row, 3, line, indent=1)
            for i in range(n_periods):
                _input_cell(ws, row, 4 + i, fmt=fmt)
            key = f"{segment}::{line}"
            row_map[key] = row
            if line == "Revenue":
                segment_revenue_rows.append(row)
            elif line == "GP":
                segment_gp_rows.append(row)
            elif line == "OP":
                segment_op_rows.append(row)
            row += 1
        # OPM (%) — formula: OP / Revenue
        _label(ws, row, 3, "OPM (%)", indent=1)
        for i in range(n_periods):
            col_letter = get_column_letter(4 + i)
            rev_r = row_map[f"{segment}::Revenue"]
            op_r = row_map[f"{segment}::OP"]
            _formula_cell(
                ws, row, 4 + i,
                f"=IFERROR({col_letter}{op_r}/{col_letter}{rev_r},0)",
                fmt="0.0%",
            )
        row += 1
        # 공백
        row += 1

    # Total — SUM of segments
    _subhdr(ws, row, 2, "Total", span=2 + n_periods)
    row += 1
    total_start = row
    for line, source_rows in [("Revenue", segment_revenue_rows),
                              ("GP", segment_gp_rows),
                              ("OP", segment_op_rows)]:
        _label(ws, row, 3, line, indent=1)
        for i in range(n_periods):
            col_letter = get_column_letter(4 + i)
            refs = ",".join(f"{col_letter}{r}" for r in source_rows)
            _formula_cell(ws, row, 4 + i, f"=SUM({refs})", fmt="#,##0.0")
        row_map[f"Total::{line}"] = row
        row += 1

    # GP margin, OP margin
    for line, num_key in [("GP margin (%)", "GP"),
                          ("OP margin (%)", "OP")]:
        _label(ws, row, 3, line, indent=1)
        for i in range(n_periods):
            col_letter = get_column_letter(4 + i)
            num_r = row_map[f"Total::{num_key}"]
            den_r = row_map[f"Total::Revenue"]
            _formula_cell(
                ws, row, 4 + i,
                f"=IFERROR({col_letter}{num_r}/{col_letter}{den_r},0)",
                fmt="0.0%",
            )
        row += 1

    return row_map


def _build_is(wb, n_periods: int, periods: list[str],
              segment_refs: dict[str, int], unit: str) -> dict[str, int]:
    """IS — Revenue·OP은 Segment_Buildup Total에서 참조, 나머지는 input."""
    ws = wb.create_sheet("IS")
    _set_widths(ws, {1: 4, 2: 32})
    for c in range(3, 3 + n_periods):
        ws.column_dimensions[get_column_letter(c)].width = 12
    _hdr(ws, 1, 2, f"💰 Income Statement ({unit})", span=1 + n_periods)
    for i, p in enumerate(periods):
        _subhdr(ws, 2, 3 + i, p)

    spec = [
        {"label": "Revenue", "kind": "formula",
         "expr": "[Segment_Buildup:Revenue]"},
        {"label": "  YoY (%)", "kind": "input", "fmt": "0.0%"},
        {"label": "(-) COGS", "kind": "input"},
        {"label": "Gross Profit", "kind": "formula",
         "expr": "[Revenue]-[(-) COGS]"},
        {"label": "GP Margin (%)", "kind": "formula",
         "expr": "IFERROR([Gross Profit]/[Revenue],0)", "fmt": "0.0%"},
        {"label": "", "kind": "blank"},
        {"label": "(-) SG&A", "kind": "input"},
        {"label": "(-) R&D", "kind": "input"},
        {"label": "(-) 기타 영업비용", "kind": "input"},
        {"label": "Operating Profit", "kind": "formula",
         "expr": "[Gross Profit]-[(-) SG&A]-[(-) R&D]-[(-) 기타 영업비용]"},
        {"label": "OP Margin (%)", "kind": "formula",
         "expr": "IFERROR([Operating Profit]/[Revenue],0)", "fmt": "0.0%"},
        {"label": "", "kind": "blank"},
        {"label": "(+) 금융수익", "kind": "input"},
        {"label": "(-) 금융비용", "kind": "input"},
        {"label": "(+) 지분법손익", "kind": "input"},
        {"label": "(+/-) 기타 영업외", "kind": "input"},
        {"label": "Pre-tax Profit (PBT)", "kind": "formula",
         "expr": "[Operating Profit]+[(+) 금융수익]-[(-) 금융비용]"
                  "+[(+) 지분법손익]+[(+/-) 기타 영업외]"},
        {"label": "(-) Income tax", "kind": "input"},
        {"label": "Net Income", "kind": "formula",
         "expr": "[Pre-tax Profit (PBT)]-[(-) Income tax]"},
        {"label": "(-) Minority interest", "kind": "input"},
        {"label": "Net Income (지배)", "kind": "formula",
         "expr": "[Net Income]-[(-) Minority interest]"},
        {"label": "", "kind": "blank"},
        {"label": "(+) D&A", "kind": "input"},
        {"label": "EBITDA", "kind": "formula",
         "expr": "[Operating Profit]+[(+) D&A]"},
        {"label": "EBITDA Margin (%)", "kind": "formula",
         "expr": "IFERROR([EBITDA]/[Revenue],0)", "fmt": "0.0%"},
        {"label": "", "kind": "blank"},
        {"label": "Shares outstanding (m)", "kind": "input", "fmt": "#,##0.0"},
        {"label": "EPS", "kind": "formula",
         "expr": "IFERROR([Net Income (지배)]/[Shares outstanding (m)]*1000,0)",
         "fmt": "#,##0"},
        {"label": "DPS", "kind": "input", "fmt": "#,##0"},
        {"label": "Payout ratio (%)", "kind": "formula",
         "expr": "IFERROR([DPS]*[Shares outstanding (m)]/1000/[Net Income (지배)],0)",
         "fmt": "0.0%"},
    ]
    return _build_sheet(ws, spec, label_col=2, data_col_start=3,
                        n_periods=n_periods,
                        cross_refs={"Segment_Buildup": {
                            "Revenue": segment_refs["Total::Revenue"],
                        }})


def _build_bs(wb, n_periods: int, periods: list[str], unit: str,
              is_refs: dict[str, int]) -> dict[str, int]:
    """BS — 합계는 자동, ROE/Debt/Equity는 IS 참조."""
    ws = wb.create_sheet("BS")
    _set_widths(ws, {1: 4, 2: 32})
    for c in range(3, 3 + n_periods):
        ws.column_dimensions[get_column_letter(c)].width = 12
    _hdr(ws, 1, 2, f"📋 Balance Sheet ({unit})", span=1 + n_periods)
    for i, p in enumerate(periods):
        _subhdr(ws, 2, 3 + i, p)

    spec = [
        {"label": "[자산]", "kind": "subheader"},
        {"label": "현금성 자산", "kind": "input"},
        {"label": "매출채권", "kind": "input"},
        {"label": "재고자산", "kind": "input"},
        {"label": "기타 유동자산", "kind": "input"},
        {"label": "유동자산 합계", "kind": "formula",
         "expr": "SUM([현금성 자산]:[기타 유동자산])"},
        {"label": "유형자산 (PP&E)", "kind": "input"},
        {"label": "무형자산", "kind": "input"},
        {"label": "투자자산", "kind": "input"},
        {"label": "기타 비유동자산", "kind": "input"},
        {"label": "비유동자산 합계", "kind": "formula",
         "expr": "SUM([유형자산 (PP&E)]:[기타 비유동자산])"},
        {"label": "자산 총계", "kind": "formula",
         "expr": "[유동자산 합계]+[비유동자산 합계]"},
        {"label": "", "kind": "blank"},
        {"label": "[부채]", "kind": "subheader"},
        {"label": "단기차입금", "kind": "input"},
        {"label": "매입채무", "kind": "input"},
        {"label": "기타 유동부채", "kind": "input"},
        {"label": "유동부채 합계", "kind": "formula",
         "expr": "SUM([단기차입금]:[기타 유동부채])"},
        {"label": "장기차입금", "kind": "input"},
        {"label": "사채", "kind": "input"},
        {"label": "기타 비유동부채", "kind": "input"},
        {"label": "비유동부채 합계", "kind": "formula",
         "expr": "SUM([장기차입금]:[기타 비유동부채])"},
        {"label": "부채 총계", "kind": "formula",
         "expr": "[유동부채 합계]+[비유동부채 합계]"},
        {"label": "", "kind": "blank"},
        {"label": "[자본]", "kind": "subheader"},
        {"label": "자본금", "kind": "input"},
        {"label": "자본잉여금", "kind": "input"},
        {"label": "이익잉여금", "kind": "input"},
        {"label": "기타 자본구성요소", "kind": "input"},
        {"label": "자본 총계", "kind": "formula",
         "expr": "SUM([자본금]:[기타 자본구성요소])"},
        {"label": "", "kind": "blank"},
        {"label": "Total debt", "kind": "formula",
         "expr": "[단기차입금]+[장기차입금]+[사채]"},
        {"label": "Net Debt (Cash)", "kind": "formula",
         "expr": "[Total debt]-[현금성 자산]"},
        {"label": "Debt/Equity (%)", "kind": "formula",
         "expr": "IFERROR([Total debt]/[자본 총계],0)", "fmt": "0.0%"},
        {"label": "ROE (%)", "kind": "formula",
         "expr": "IFERROR([IS:Net Income (지배)]/[자본 총계],0)", "fmt": "0.0%"},
        {"label": "ROA (%)", "kind": "formula",
         "expr": "IFERROR([IS:Net Income (지배)]/[자산 총계],0)", "fmt": "0.0%"},
    ]
    return _build_sheet(ws, spec, label_col=2, data_col_start=3,
                        n_periods=n_periods,
                        cross_refs={"IS": is_refs})


def _build_cf(wb, n_periods: int, periods: list[str], unit: str,
              is_refs: dict[str, int]) -> dict[str, int]:
    """CF — Net Income은 IS 참조, 합계는 자동."""
    ws = wb.create_sheet("CF")
    _set_widths(ws, {1: 4, 2: 32})
    for c in range(3, 3 + n_periods):
        ws.column_dimensions[get_column_letter(c)].width = 12
    _hdr(ws, 1, 2, f"💵 Cash Flow Statement ({unit})", span=1 + n_periods)
    for i, p in enumerate(periods):
        _subhdr(ws, 2, 3 + i, p)

    spec = [
        {"label": "[영업활동]", "kind": "subheader"},
        {"label": "Net Income", "kind": "formula",
         "expr": "[IS:Net Income]"},
        {"label": "(+) D&A", "kind": "formula", "expr": "[IS:(+) D&A]"},
        {"label": "(+/-) 운전자본 변동", "kind": "input"},
        {"label": "(+/-) 기타 (비현금)", "kind": "input"},
        {"label": "영업활동 CF", "kind": "formula",
         "expr": "[Net Income]+[(+) D&A]+[(+/-) 운전자본 변동]+[(+/-) 기타 (비현금)]"},
        {"label": "", "kind": "blank"},
        {"label": "[투자활동]", "kind": "subheader"},
        {"label": "(-) Capex", "kind": "input"},
        {"label": "(-) 무형자산 취득", "kind": "input"},
        {"label": "(+/-) 투자자산 매매", "kind": "input"},
        {"label": "(+/-) 기타 투자", "kind": "input"},
        {"label": "투자활동 CF", "kind": "formula",
         "expr": "[(-) Capex]+[(-) 무형자산 취득]+[(+/-) 투자자산 매매]+[(+/-) 기타 투자]"},
        {"label": "", "kind": "blank"},
        {"label": "[재무활동]", "kind": "subheader"},
        {"label": "(+/-) 차입금 증감", "kind": "input"},
        {"label": "(-) 배당금", "kind": "input"},
        {"label": "(+/-) 자기주식", "kind": "input"},
        {"label": "(+/-) 기타 재무", "kind": "input"},
        {"label": "재무활동 CF", "kind": "formula",
         "expr": "[(+/-) 차입금 증감]+[(-) 배당금]+[(+/-) 자기주식]+[(+/-) 기타 재무]"},
        {"label": "", "kind": "blank"},
        {"label": "Net Cash Change", "kind": "formula",
         "expr": "[영업활동 CF]+[투자활동 CF]+[재무활동 CF]"},
        {"label": "Cash, beginning", "kind": "input"},
        {"label": "Cash, ending", "kind": "formula",
         "expr": "[Cash, beginning]+[Net Cash Change]"},
        {"label": "", "kind": "blank"},
        {"label": "Free Cash Flow (영업 - Capex)", "kind": "formula",
         "expr": "[영업활동 CF]+[(-) Capex]"},
        {"label": "FCF margin (%)", "kind": "formula",
         "expr": "IFERROR([Free Cash Flow (영업 - Capex)]/[IS:Revenue],0)",
         "fmt": "0.0%"},
    ]
    return _build_sheet(ws, spec, label_col=2, data_col_start=3,
                        n_periods=n_periods,
                        cross_refs={"IS": is_refs})


def _build_valuation_band(wb, template: IndustryTemplate, n_periods: int,
                           periods: list[str], hist_n: int) -> dict[str, int]:
    """Valuation Band — multiple input 후 historical 구간으로 avg/median/σ."""
    ws = wb.create_sheet("Valuation_Band")
    _set_widths(ws, {1: 4, 2: 24})
    for c in range(3, 3 + n_periods):
        ws.column_dimensions[get_column_letter(c)].width = 11

    _hdr(ws, 1, 2, "📈 Valuation Band — Historical + projection",
         span=1 + n_periods)
    for i, p in enumerate(periods):
        _subhdr(ws, 2, 3 + i, p)

    multiple_rows: dict[str, int] = {}
    row = 3
    for m in template.valuation_multiples:
        _label(ws, row, 2, m)
        for i in range(n_periods):
            _input_cell(ws, row, 3 + i, fmt="#,##0.00")
        multiple_rows[m] = row
        row += 1

    row += 1
    _subhdr(ws, row, 2, "Statistical bands (historical 구간)",
            span=1 + n_periods)
    row += 1
    # Stats — AVERAGE/MEDIAN/STDEV/MIN/MAX over historical columns (first hist_n)
    hist_col_start = 3
    hist_col_end = 3 + hist_n - 1
    for stat_label, stat_fn in [
        (f"Avg ({hist_n}y)", "AVERAGE"),
        (f"Median ({hist_n}y)", "MEDIAN"),
        ("1 σ", "STDEV"),
        ("Min", "MIN"),
        ("Max", "MAX"),
    ]:
        _label(ws, row, 2, stat_label)
        for m in template.valuation_multiples:
            # For each multiple, compute stat over its historical range
            pass
        # 각 multiple마다 한 셀씩 stat 계산 — 시각화 위해 첫 multiple만 (단순화):
        for i, m in enumerate(template.valuation_multiples):
            if i < n_periods:
                mr = multiple_rows[m]
                col_letter = get_column_letter(3 + i)
                start = f"{get_column_letter(hist_col_start)}{mr}"
                end = f"{get_column_letter(hist_col_end)}{mr}"
                _formula_cell(ws, row, 3 + i, f"={stat_fn}({start}:{end})",
                              fmt="#,##0.00")
        row += 1

    row += 1
    _subhdr(ws, row, 2, "Implied target price", span=1 + n_periods)
    row += 1
    impl = {}
    for line in ("Target multiple", "EPS or EBITDA per share", "Implied price"):
        _label(ws, row, 2, line)
        for i in range(n_periods):
            _input_cell(ws, row, 3 + i, fmt="#,##0.00")
        impl[line] = row
        row += 1
    # Upside = Implied / Current - 1 (Current price는 사용자가 첫 셀에 입력)
    _label(ws, row, 2, "Current price")
    for i in range(n_periods):
        _input_cell(ws, row, 3 + i, fmt="#,##0")
    impl["Current price"] = row
    row += 1
    _label(ws, row, 2, "Upside (%)")
    for i in range(n_periods):
        col_letter = get_column_letter(3 + i)
        _formula_cell(
            ws, row, 3 + i,
            f"=IFERROR({col_letter}{impl['Implied price']}/"
            f"{col_letter}{impl['Current price']}-1,0)",
            fmt="0.0%",
        )
    row += 1

    return multiple_rows


def _build_sotp(wb, template: IndustryTemplate):
    """SOTP — Segment × Multiple = EV → Equity → Target."""
    ws = wb.create_sheet("SOTP")
    _set_widths(ws, {1: 4, 2: 22, 3: 14, 4: 14, 5: 14, 6: 14, 7: 14, 8: 16})
    _hdr(ws, 1, 2, "💎 Sum-of-the-parts Valuation", span=7)
    headers = ["Segment", "Metric", "Value", "Multiple", "EV",
               "Stake (%)", "Attributable EV"]
    for i, h in enumerate(headers):
        _subhdr(ws, 2, 2 + i, h)

    row = 3
    seg_ev_rows: list[int] = []
    for seg in template.segments:
        _label(ws, row, 2, seg)
        # Metric / Value / Multiple / Stake — input
        _input_cell(ws, row, 3, fmt="@")   # Metric (텍스트)
        _input_cell(ws, row, 4, fmt="#,##0.0")   # Value
        _input_cell(ws, row, 5, fmt="#,##0.0")   # Multiple
        # EV = Value × Multiple
        _formula_cell(ws, row, 6, f"=D{row}*E{row}", fmt="#,##0.0")
        _input_cell(ws, row, 7, value=100, fmt="0.0%")   # Stake
        # Attributable EV = EV × Stake
        _formula_cell(ws, row, 8, f"=F{row}*G{row}", fmt="#,##0.0")
        seg_ev_rows.append(row)
        row += 1

    row += 1
    _subhdr(ws, row, 2, "Operating EV (합계)", span=7)
    op_ev_row = row + 1
    _label(ws, op_ev_row, 2, "Sum of attributable EV")
    refs = ",".join(f"H{r}" for r in seg_ev_rows)
    _formula_cell(ws, op_ev_row, 8, f"=SUM({refs})", fmt="#,##0.0")
    row = op_ev_row + 1

    # Net cash + investments - minority → Equity → Target
    line_rows = {}
    for line in ("(+) Net cash (BS)", "(+) Investments at market",
                 "(-) Minority interest", "Equity value",
                 "Total shares (m)", "Target price (₩ or $)",
                 "Current price", "Upside (%)"):
        _label(ws, row, 2, line)
        if line == "Equity value":
            _formula_cell(
                ws, row, 8,
                f"=H{op_ev_row}+H{line_rows['(+) Net cash (BS)']}"
                f"+H{line_rows['(+) Investments at market']}"
                f"-H{line_rows['(-) Minority interest']}",
                fmt="#,##0.0",
            )
        elif line == "Target price (₩ or $)":
            _formula_cell(
                ws, row, 8,
                f"=IFERROR(H{line_rows['Equity value']}*1000"
                f"/H{line_rows['Total shares (m)']},0)",
                fmt="#,##0",
            )
        elif line == "Upside (%)":
            _formula_cell(
                ws, row, 8,
                f"=IFERROR(H{line_rows['Target price (₩ or $)']}"
                f"/H{line_rows['Current price']}-1,0)",
                fmt="0.0%",
            )
        else:
            _input_cell(ws, row, 8, fmt="#,##0.0")
        line_rows[line] = row
        row += 1


def _build_summary(wb, n_periods: int, periods: list[str], unit: str,
                   is_refs: dict[str, int], bs_refs: dict[str, int],
                   cf_refs: dict[str, int]):
    """Summary — 모든 KPI를 IS/BS/CF에서 참조."""
    ws = wb.create_sheet("Summary", 1)
    _set_widths(ws, {1: 4, 2: 28})
    for c in range(3, 3 + n_periods):
        ws.column_dimensions[get_column_letter(c)].width = 11

    _hdr(ws, 1, 2, f"📊 Summary — Key financials & KPI ({unit})",
         span=1 + n_periods)
    for i, p in enumerate(periods):
        _subhdr(ws, 2, 3 + i, p)

    cross = {"IS": is_refs, "BS": bs_refs, "CF": cf_refs}
    spec = [
        {"label": "[손익]", "kind": "subheader"},
        {"label": "Revenue", "kind": "formula", "expr": "[IS:Revenue]"},
        {"label": "  YoY (%)", "kind": "input", "fmt": "0.0%"},
        {"label": "Gross Profit", "kind": "formula", "expr": "[IS:Gross Profit]"},
        {"label": "Operating Profit", "kind": "formula", "expr": "[IS:Operating Profit]"},
        {"label": "  OPM (%)", "kind": "formula", "expr": "[IS:OP Margin (%)]",
         "fmt": "0.0%"},
        {"label": "EBITDA", "kind": "formula", "expr": "[IS:EBITDA]"},
        {"label": "  EBITDA Margin (%)", "kind": "formula",
         "expr": "[IS:EBITDA Margin (%)]", "fmt": "0.0%"},
        {"label": "Net Income (지배)", "kind": "formula",
         "expr": "[IS:Net Income (지배)]"},
        {"label": "EPS", "kind": "formula", "expr": "[IS:EPS]", "fmt": "#,##0"},
        {"label": "DPS", "kind": "formula", "expr": "[IS:DPS]", "fmt": "#,##0"},
        {"label": "", "kind": "blank"},
        {"label": "[현금흐름·재무건전성]", "kind": "subheader"},
        {"label": "OCF", "kind": "formula", "expr": "[CF:영업활동 CF]"},
        {"label": "Capex", "kind": "formula", "expr": "[CF:(-) Capex]"},
        {"label": "FCF", "kind": "formula",
         "expr": "[CF:Free Cash Flow (영업 - Capex)]"},
        {"label": "Net Debt (Cash)", "kind": "formula",
         "expr": "[BS:Net Debt (Cash)]"},
        {"label": "ROE (%)", "kind": "formula", "expr": "[BS:ROE (%)]",
         "fmt": "0.0%"},
        {"label": "Debt/Equity (%)", "kind": "formula",
         "expr": "[BS:Debt/Equity (%)]", "fmt": "0.0%"},
        {"label": "", "kind": "blank"},
        {"label": "[밸류에이션]", "kind": "subheader"},
        {"label": "Market cap", "kind": "input"},
        {"label": "EV (= MC + Net Debt)", "kind": "formula",
         "expr": "[Market cap]+[Net Debt (Cash)]"},
        {"label": "P/E", "kind": "formula",
         "expr": "IFERROR([Market cap]/[Net Income (지배)],0)", "fmt": "#,##0.00"},
        {"label": "EV/EBITDA", "kind": "formula",
         "expr": "IFERROR([EV (= MC + Net Debt)]/[EBITDA],0)",
         "fmt": "#,##0.00"},
        {"label": "EV/Sales", "kind": "formula",
         "expr": "IFERROR([EV (= MC + Net Debt)]/[Revenue],0)",
         "fmt": "#,##0.00"},
        {"label": "Div yield (%)", "kind": "formula",
         "expr": "IFERROR([DPS]*1000/[Market cap],0)", "fmt": "0.0%"},
    ]
    _build_sheet(ws, spec, label_col=2, data_col_start=3,
                 n_periods=n_periods, cross_refs=cross)


def _build_notes(wb, template: IndustryTemplate):
    ws = wb.create_sheet("Notes")
    _set_widths(ws, {1: 4, 2: 90})
    _hdr(ws, 1, 2, "📝 Notes & Assumptions")
    row = 3
    blocks = [
        ("Industry template:", f"{template.label} ({template.code})"),
        ("Segments:", " · ".join(template.segments)),
        ("Key drivers:",
         "\n".join(f"  - [{s}] {d} ({u})" for s, d, u in template.drivers)),
        ("Notes:", template.notes),
        ("",
         "Color: 🟡 input · 🟢 formula · 일반 = label\n"
         "Workflow: Drivers → Segment_Buildup → IS/BS/CF (자동 chain) → "
         "Valuation_Band/SOTP → Summary"),
    ]
    for title, body in blocks:
        if title:
            c = ws.cell(row=row, column=2, value=title)
            c.font = SUBHEADER_FONT
            row += 1
        if body:
            c = ws.cell(row=row, column=2, value=body)
            c.font = DATA_FONT
            c.alignment = WRAP
            ws.row_dimensions[row].height = max(20, body.count("\n") * 18 + 24)
            row += 1
        row += 1


# ─────────────────────── 통합 ───────────────────────
def generate_model_xlsx(
    ticker: str,
    name: str,
    industry: str = "general",
    hist_years: list[int] | None = None,
    proj_years: list[int] | None = None,
    unit: str = "₩bn",
    analyst: str = "",
    out_path: str | None = None,
) -> str:
    """sell-side 빈 모델 xlsx + formula chain 자동 wiring 생성."""
    import datetime as _dt
    if hist_years is None:
        cy = _dt.date.today().year
        hist_years = [cy - 5, cy - 4, cy - 3, cy - 2, cy - 1]
    if proj_years is None:
        cy = _dt.date.today().year
        proj_years = [cy, cy + 1, cy + 2]

    template = INDUSTRY_TEMPLATES.get(industry, INDUSTRY_TEMPLATES["general"])
    periods = _periods(hist_years, proj_years)
    n_periods = len(periods)
    hist_n = len(hist_years)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 빌드 순서 — Segment_Buildup이 IS의 cross-ref 소스이므로 먼저
    _build_cover(wb, ticker, name, template.label, analyst)
    _build_drivers(wb, template, n_periods, periods)
    segment_refs = _build_segment(wb, template, n_periods, periods)
    is_refs = _build_is(wb, n_periods, periods, segment_refs, unit)
    bs_refs = _build_bs(wb, n_periods, periods, unit, is_refs)
    cf_refs = _build_cf(wb, n_periods, periods, unit, is_refs)
    _build_valuation_band(wb, template, n_periods, periods, hist_n)
    _build_sotp(wb, template)
    _build_summary(wb, n_periods, periods, unit, is_refs, bs_refs, cf_refs)
    _build_notes(wb, template)

    # 시트 순서 정리: Cover / Summary / IS / BS / CF / Drivers / Segment_Buildup /
    #                Valuation_Band / SOTP / Notes
    order = ["Cover", "Summary", "IS", "BS", "CF", "Drivers",
             "Segment_Buildup", "Valuation_Band", "SOTP", "Notes"]
    sheet_order = [wb[s] for s in order if s in wb.sheetnames]
    other_sheets = [s for s in wb.worksheets if s.title not in order]
    wb._sheets = sheet_order + other_sheets

    if out_path is None:
        out_path = tempfile.NamedTemporaryFile(
            prefix=f"model_{ticker.replace('.', '_')}_",
            suffix=".xlsx", delete=False,
        ).name
    wb.save(out_path)
    log.info("model xlsx saved: %s (%d sheets)", out_path, len(wb.sheetnames))
    return out_path


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    path = generate_model_xlsx(
        ticker="064960.KS", name="SNT Motiv",
        industry="auto_parts", unit="₩bn",
        analyst="기계/자동차",
        out_path=r"C:\Users\srkwn\Downloads\SNT_Motiv_model_v2.xlsx",
    )
    print(f"\n✓ {path}")
    print(f"  사이즈: {Path(path).stat().st_size:,} bytes")
